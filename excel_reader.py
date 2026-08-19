"""
excel_reader.py — Extrae y normaliza los datos de la bitácora de crédito Excel.

El archivo "BITACORA CUPOS DE CREDITO Y MAS 2026.xlsx" contiene:
  - Hoja1: Registros principales de solicitudes de crédito (200 filas)
  - N Cred.: Solicitudes nuevas/pendientes (102 filas)
  - Estudio de EF / Hoja3: Análisis financiero y estados financieros

Columnas relevantes de Hoja1:
  - NOMBRE CLIENTE, CEDULA NIT
  - CREDITO ACTUAL, MONTO A SOLICITAR, CUPO INICIAL, CREDITO APROBADO
  - PROMEDIO (compras), COMPRA MINIMA, COMPRA MAXIMA, NUMERO DE COMPRAS
  - AÑO DEL DATO DE COMPRAS, PROMEDIO DE PAGO (DIAS)
  - CALIFIC PROMEDIO DATA CREDITO, CONSULTA ULTIMOS 6 MESES SECTOR REAL
  - PRESENTA MORA, PRESENTA CARTERA CASTIGADA
  - APROBACION, OBSERVACIONES
"""
from __future__ import annotations

import re
import time
import logging
import threading
from datetime import datetime, date
from pathlib import Path
from typing import Any

log = logging.getLogger("verifydata.excel")

EXCEL_PATH = Path(__file__).parent / "creditos_personas" / "BITACORA CUPOS DE CREDITO Y MAS 2026.xlsx"

# ── Caché con TTL (evita re-leer el Excel en cada request) ──
_excel_cache: dict[str, Any] | None = None
_excel_cache_time: float = 0.0
_excel_cache_lock = threading.Lock()
EXCEL_CACHE_TTL = 300  # 5 minutos

# Mapeo nombre de columna del Excel → key normalizada
COLUMN_MAP_HOJA1 = {
    "FECHA LLEGADA DOC FISICO": "fecha_llegada_doc",
    "DOC EN OFICINA": "doc_en_oficina",
    "FECHA SOLICITUD": "fecha_solicitud",
    "FECHA RESPUESTA": "fecha_respuesta",
    "TIPO DE SOLICITUD": "tipo_solicitud",
    "EJECUTIVO": "ejecutivo",
    "NOMBRE CLIENTE": "nombre_cliente",
    "CEDULA NIT": "cedula_nit",
    "SUCURSAL": "sucursal",
    "FORMATO ENTREVISTA": "formato_entrevista",
    "SOL. DE CREDITO": "solicitud_credito",
    "PAGARE": "pagare",
    "CARTA DE INSTRUCC": "carta_instrucciones",
    "CREDITO ACTUAL": "credito_actual",
    "MONTO A SOLICITAR": "monto_solicitar",
    "RUT": "rut",
    "CAM DE CCIO": "camara_comercio",
    "CEDULA RL": "cedula_rl",
    "ESTADOS FINANC": "estados_financieros",
    "DECLARAC DE RENTA": "declaracion_renta",
    "FACTURA O CERTIFICACION 1": "factura_1",
    "VALOR 1": "valor_1",
    "FACTURA O CERTIFICACION 2": "factura_2",
    "VALOR 2": "valor_2",
    "FACTURA O CERTIFICACION 3": "factura_3",
    "VALOR 3": "valor_3",
    "FACTURA O CERTIFICACION 4": "factura_4",
    "VALOR 4": "valor_4",
    "FACTURA O CERTIFICACION 5": "factura_5",
    "VALOR 5": "valor_5",
    "PROMEDIO": "promedio_compras",
    "COMPRA MINIMA": "compra_minima",
    "COMPRA MAXIMA": "compra_maxima",
    "NUMERO DE COMPRAS": "numero_compras",
    "AÑO DEL DATO DE COMPRAS": "ano_dato_compras",
    "PROMEDIO DE PAGO (DIAS)": "promedio_pago_dias",
    "CALIFIC PROMEDIO DATA CREDITO": "calificacion_datacredito",
    "CUPO INICIAL": "cupo_inicial",
    "CONSULTA ULTIMOS 6 MESES SECTOR REAL": "consultas_6m_sector_real",
    "PRESENTA MORA": "presenta_mora",
    "PRESENTA  CARTERA CASTIGADA": "presenta_cartera_castigada",
    "APROBACION": "aprobacion",
    "CREDITO APROBADO": "credito_aprobado",
    "OBSERVACIONES": "observaciones",
}

COLUMN_MAP_N_CRED = {
    "FECHA LLEGADA DOC FISICO": "fecha_llegada_doc",
    "DÍAS RECIBIDO DCTOS FISICOS": "dias_recibido_docs",
    "DOC EN OFICINA": "doc_en_oficina",
    "DÍAS EN LOS QUE SE DIO RESPUESTA": "dias_respuesta",
    "FECHA SOLICITUD": "fecha_solicitud",
    "FECHA RESPUESTA": "fecha_respuesta",
    "CLIENTE NUEVO": "cliente_nuevo",
    "TIPO DE SOLICITUD": "tipo_solicitud",
    "EJECUTIVO": "ejecutivo",
    "NOMBRE CLIENTE": "nombre_cliente",
    "CEDULA NIT": "cedula_nit",
    "SUCURSAL": "sucursal",
    "FORMATO ENTREVISTA": "formato_entrevista",
    "SOL. DE CREDITO/PAGARÉ/CARTA DE INSTRUCCIONES": "docs_credito",
    "CREDITO ACTUAL": "credito_actual",
    "MONTO A SOLICITAR": "monto_solicitar",
    "RUT": "rut",
    "CAM DE CCIO": "camara_comercio",
    "CEDULA RL": "cedula_rl",
    "ESTADOS FINANC": "estados_financieros",
    "DECLARAC DE RENTA": "declaracion_renta",
    "FACTURA O CERTIFICACION 1": "factura_1",
    "VALOR 1": "valor_1",
    "FACTURA O CERTIFICACION 2": "factura_2",
    "VALOR 2": "valor_2",
    "FACTURA O CERTIFICACION 3": "factura_3",
    "VALOR 3": "valor_3",
    "FACTURA O CERTIFICACION 4": "factura_4",
    "VALOR 4": "valor_4",
    "FACTURA O CERTIFICACION 5": "factura_5",
    "VALOR 5": "valor_5",
    "PROMEDIO": "promedio_compras",
    "COMPRA MINIMA": "compra_minima",
    "COMPRA MAXIMA": "compra_maxima",
    "NUMERO DE COMPRAS": "numero_compras",
    "AÑO DEL DATO DE COMPRAS": "ano_dato_compras",
    "PROMEDIO DE PAGO (DIAS)": "promedio_pago_dias",
    "CALIFIC PROMEDIO DATA CREDITO": "calificacion_datacredito",
    "CUPO INICIAL": "cupo_inicial",
    "CONSULTA ULTIMOS 6 MESES SECTOR REAL": "consultas_6m_sector_real",
    "PRESENTA MORA": "presenta_mora",
    "PRESENTA  CARTERA CASTIGADA": "presenta_cartera_castigada",
    "RESPUESTA": "respuesta",
    "MONTO DE CREDITO APROBADO": "monto_credito_aprobado",
    "OBSERVACIONES": "observaciones",
}


def _parse_number(val: Any) -> float | None:
    """Convierte un valor de celda a número o None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "").replace("%", "")
    if not s or s.lower() in ("none", "n/a", "-", "no", "si", "sí"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_bool(val: Any) -> bool | None:
    """Convierte un valor a booleano."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    if s in ("SI", "SÍ", "YES", "TRUE", "1", "Y"):
        return True
    if s in ("NO", "FALSE", "0", "N"):
        return False
    return None


def _parse_date(val: Any) -> str | None:
    """Convierte una fecha de Excel a string ISO."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ("none", ""):
        return None
    # Intentar parsear formatos comunes
    for fmt in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y",
        "%m/%d/%Y", "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Si solo es año
    m = re.match(r"^(\d{4})$", s)
    if m:
        return s
    return s


def _clean_text(val: Any) -> str:
    """Limpia texto de celda."""
    if val is None:
        return ""
    return str(val).strip()


def _normalize_cedula(raw: str) -> str:
    """Normaliza una cédula/NIT: solo dígitos."""
    return re.sub(r"\D", "", raw)


def _normalize_name(raw: str) -> str:
    """Normaliza nombre: uppercase, sin espacios extra, sin tildes."""
    import unicodedata
    s = raw.strip().upper()
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    s = re.sub(r"\s+", " ", s)
    return s


def read_hoja1(path: Path | None = None) -> list[dict[str, Any]]:
    """Lee Hoja1 (solicitudes de crédito históricas)."""
    return _read_sheet(path, "Hoja1", COLUMN_MAP_HOJA1)


def read_n_cred(path: Path | None = None) -> list[dict[str, Any]]:
    """Lee Hoja 'N Cred.' (solicitudes nuevas)."""
    return _read_sheet(path, "N Cred.", COLUMN_MAP_N_CRED)


def _read_sheet(
    path: Path | None, sheet_name: str, col_map: dict[str, str]
) -> list[dict[str, Any]]:
    """Lee una hoja del Excel y devuelve lista de diccionarios normalizados."""
    fp = path or EXCEL_PATH
    if not fp.exists():
        log.error("Excel no encontrado: %s", fp)
        return []

    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl no instalado. pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(fp, data_only=True)
    if sheet_name not in wb.sheetnames:
        log.warning("Hoja %r no encontrada en el Excel", sheet_name)
        return []

    ws = wb[sheet_name]

    # Leer headers
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=col).value or "").strip()
        headers.append(h)

    # Mapear índices de columna → key normalizada
    col_index: dict[int, str] = {}
    for i, h in enumerate(headers):
        if h in col_map:
            col_index[i] = col_map[h]

    # Leer filas
    records = []
    for row in range(2, ws.max_row + 1):
        record: dict[str, Any] = {}
        for col_i, key in col_index.items():
            raw = ws.cell(row=row, column=col_i + 1).value
            record[key] = raw

        # Saltar filas completamente vacías
        if not any(
            v is not None and str(v).strip() not in ("", "None")
            for v in record.values()
        ):
            continue

        # Normalizar
        if "cedula_nit" in record:
            record["cedula_nit"] = _normalize_cedula(
                str(record["cedula_nit"] or "")
            )
        if "nombre_cliente" in record:
            record["nombre_cliente_normalizado"] = _normalize_name(
                str(record["nombre_cliente"] or "")
            )

        # Parsear fechas
        for fk in ("fecha_solicitud", "fecha_respuesta", "fecha_llegada_doc"):
            if fk in record:
                record[fk] = _parse_date(record[fk])

        # Parsear montos
        for mk in (
            "credito_actual", "monto_solicitar", "cupo_inicial",
            "credito_aprobado", "monto_credito_aprobado",
            "promedio_compras", "compra_minima", "compra_maxima",
            "valor_1", "valor_2", "valor_3", "valor_4", "valor_5",
        ):
            if mk in record:
                record[mk] = _parse_number(record[mk])

        # Parsear número de compras y año
        if "numero_compras" in record:
            record["numero_compras"] = (
                int(_parse_number(record["numero_compras"]) or 0)
            )
        if "ano_dato_compras" in record:
            v = record["ano_dato_compras"]
            record["ano_dato_compras"] = (
                int(v) if v is not None and str(v).isdigit() else None
            )

        # Parsear días promedio de pago
        if "promedio_pago_dias" in record:
            record["promedio_pago_dias"] = _parse_number(
                record["promedio_pago_dias"]
            )

        # Parsear calificación datacrédito
        if "calificacion_datacredito" in record:
            record["calificacion_datacredito"] = _parse_number(
                record["calificacion_datacredito"]
            )

        # Parsear booleanos
        for bk in (
            "doc_en_oficina", "formato_entrevista",
            "solicitud_credito", "pagare", "carta_instrucciones",
            "rut", "camara_comercio", "cedula_rl",
            "estados_financieros", "declaracion_renta",
            "presenta_mora", "presenta_cartera_castigada",
            "aprobacion",
        ):
            if bk in record:
                record[bk] = _parse_bool(record[bk])

        # Parsear docs_credito unificado (N Cred.)
        if "docs_credito" in record:
            record["docs_credito"] = _parse_bool(record["docs_credito"])

        records.append(record)

    wb.close()
    log.info("Excel %s: %d registros leídos", sheet_name, len(records))
    return records


def read_all(path: Path | None = None) -> dict[str, Any]:
    """Lee todas las hojas relevantes y consolida (con caché 5 min)."""
    global _excel_cache, _excel_cache_time
    now = time.time()
    with _excel_cache_lock:
        if _excel_cache and (now - _excel_cache_time) < EXCEL_CACHE_TTL:
            return _excel_cache

    hoja1 = read_hoja1(path)
    n_cred = read_n_cred(path)

    # Consolidar todos los registros únicos por cédula
    all_records = hoja1 + n_cred
    unique_cedulas: dict[str, dict[str, Any]] = {}
    for r in all_records:
        ced = r.get("cedula_nit", "")
        if not ced or len(ced) < 5:
            continue
        if ced not in unique_cedulas:
            unique_cedulas[ced] = r
        else:
            # Merge: la segunda fuente complementa la primera
            existing = unique_cedulas[ced]
            for k, v in r.items():
                if v is not None and (
                    k not in existing or existing[k] is None
                ):
                    existing[k] = v

    result = {
        "total_registros_hoja1": len(hoja1),
        "total_registros_n_cred": len(n_cred),
        "total_clientes_unicos": len(unique_cedulas),
        "clientes": list(unique_cedulas.values()),
        "hoja1": hoja1,
        "n_cred": n_cred,
    }
    with _excel_cache_lock:
        _excel_cache = result
        _excel_cache_time = time.time()
    return result


def get_client_by_cedula(
    cedula: str, path: Path | None = None
) -> dict[str, Any] | None:
    """Busca un cliente por cédula/NIT en el Excel."""
    cedula = _normalize_cedula(cedula)
    all_data = read_all(path)
    for c in all_data["clientes"]:
        if c.get("cedula_nit") == cedula:
            return c
    return None


def get_client_by_name(
    name: str, path: Path | None = None
) -> list[dict[str, Any]]:
    """Busca clientes por nombre (búsqueda parcial)."""
    name_norm = _normalize_name(name)
    all_data = read_all(path)
    results = []
    for c in all_data["clientes"]:
        cname = c.get("nombre_cliente_normalizado", "")
        if name_norm in cname:
            results.append(c)
    return results


def get_all_cedulas(path: Path | None = None) -> list[str]:
    """Devuelve todas las cédulas/NIT únicos del Excel."""
    all_data = read_all(path)
    return [c["cedula_nit"] for c in all_data["clientes"] if c.get("cedula_nit")]
